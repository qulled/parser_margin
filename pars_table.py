from logging.handlers import RotatingFileHandler
from dotenv import load_dotenv
from googleapiclient import discovery
from google.oauth2 import service_account
from googleapiclient.discovery import build
import logging
import os
import datetime as dt
import json

import openpyxl
import warnings

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
log_dir = os.path.join(BASE_DIR, 'logs/')
log_file = os.path.join(BASE_DIR, 'logs/parser_marjin.log')
console_handler = logging.StreamHandler()
file_handler = RotatingFileHandler(
    log_file,
    maxBytes=100000,
    backupCount=3,
    encoding='utf-8'
)
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s, %(levelname)s, %(message)s',
    handlers=(
        file_handler,
        console_handler
    )
)

SCOPES = ['https://www.googleapis.com/auth/spreadsheets']
CREDENTIALS_FILE = 'credentials_service.json'
credentials = service_account.Credentials.from_service_account_file(CREDENTIALS_FILE)
service = discovery.build('sheets', 'v4', credentials=credentials)
START_POSITION_FOR_PLACE = 2

dotenv_path = os.path.join(os.path.dirname(__file__), '.env')

if os.path.exists(dotenv_path):
    load_dotenv(dotenv_path)
load_dotenv('.env ')
SPREADSHEET_ID = os.getenv('SPREADSHEET_ID')


def list_articles(employees_sheet):
    list_articles = []
    for x in range(2, employees_sheet.max_row + 1):
        if employees_sheet.cell(row=x, column=6).value not in list_articles and employees_sheet.cell(row=x,
                                                                                                     column=30).value != 'Сторно продаж':
            list_articles.append(employees_sheet.cell(row=x, column=6).value)
    return list_articles


def dict_count(employees_sheet):
    dict_count = {}
    for x in range(2, employees_sheet.max_row + 1):
        if employees_sheet.cell(row=x, column=6).value not in dict_count and \
                employees_sheet.cell(row=x, column=31).value == 'Продажа':
            dict_count[employees_sheet.cell(row=x, column=6).value] = 1
        elif employees_sheet.cell(row=x, column=6).value in dict_count and \
                employees_sheet.cell(row=x, column=31).value == 'Продажа':
            dict_count[employees_sheet.cell(row=x, column=6).value] += 1
    return dict_count


def dict_count_refund(employees_sheet):
    dict_count_refund = {}
    for x in range(2, employees_sheet.max_row + 1):
        if employees_sheet.cell(row=x, column=6).value not in dict_count_refund and \
                employees_sheet.cell(row=x, column=31).value == 'Возврат':
            dict_count_refund[employees_sheet.cell(row=x, column=6).value] = 1
        elif employees_sheet.cell(row=x, column=6).value in dict_count_refund and \
                employees_sheet.cell(row=x, column=31).value == 'Возврат':
            dict_count_refund[employees_sheet.cell(row=x, column=6).value] += 1
    return dict_count_refund


def dict_WB_sell(employees_sheet):
    dict_WB_sell = {}
    for x in range(2, employees_sheet.max_row + 1):
        if employees_sheet.cell(row=x, column=6).value not in dict_WB_sell and \
                employees_sheet.cell(row=x, column=13).value == 'Продажа':
            dict_WB_sell[employees_sheet.cell(row=x, column=6).value] = \
                float(employees_sheet.cell(row=x, column=16).value)
        elif employees_sheet.cell(row=x, column=6).value in dict_WB_sell and \
                employees_sheet.cell(row=x, column=13).value == 'Продажа':
            dict_WB_sell[employees_sheet.cell(row=x, column=6).value] += \
                float(employees_sheet.cell(row=x, column=16).value)
        if employees_sheet.cell(row=x, column=6).value not in dict_WB_sell and \
                employees_sheet.cell(row=x, column=13).value == 'Возврат':
            dict_WB_sell[employees_sheet.cell(row=x, column=6).value] = \
                float(-employees_sheet.cell(row=x, column=16).value)
        elif employees_sheet.cell(row=x, column=6).value in dict_WB_sell and \
                employees_sheet.cell(row=x, column=13).value == 'Возврат':
            dict_WB_sell[employees_sheet.cell(row=x, column=6).value] -= \
                float(employees_sheet.cell(row=x, column=16).value)
    return dict_WB_sell


def dict_logistics(employees_sheet):
    dict_logistics = {}
    for x in range(2, employees_sheet.max_row + 1):
        if employees_sheet.cell(row=x, column=6).value not in dict_logistics and \
                employees_sheet.cell(row=x, column=31).value == 'Логистика':
            dict_logistics[employees_sheet.cell(row=x, column=6).value] = \
                employees_sheet.cell(row=x, column=39).value
        elif employees_sheet.cell(row=x, column=6).value in dict_logistics and \
                employees_sheet.cell(row=x, column=31).value == 'Логистика':
            dict_logistics[employees_sheet.cell(row=x, column=6).value] += \
                employees_sheet.cell(row=x, column=39).value
    return dict_logistics


def dict_get_seller(employees_sheet):
    dict_get_seller = {}
    for x in range(2, employees_sheet.max_row + 1):
        if employees_sheet.cell(row=x, column=6).value not in dict_get_seller and \
                employees_sheet.cell(row=x, column=13).value == 'Продажа':
            dict_get_seller[employees_sheet.cell(row=x, column=6).value] = \
                float(employees_sheet.cell(row=x, column=36).value)
        elif employees_sheet.cell(row=x, column=6).value in dict_get_seller and \
                employees_sheet.cell(row=x, column=13).value == 'Продажа':
            dict_get_seller[employees_sheet.cell(row=x, column=6).value] += \
                float(employees_sheet.cell(row=x, column=36).value)
        if employees_sheet.cell(row=x, column=6).value not in dict_get_seller and \
                employees_sheet.cell(row=x, column=13).value == 'Возврат':
            dict_get_seller[employees_sheet.cell(row=x, column=6).value] = \
                float(-employees_sheet.cell(row=x, column=36).value)
        elif employees_sheet.cell(row=x, column=6).value in dict_get_seller and \
                employees_sheet.cell(row=x, column=13).value == 'Возврат':
            dict_get_seller[employees_sheet.cell(row=x, column=6).value] -= \
                float(employees_sheet.cell(row=x, column=36).value)
    return dict_get_seller


def convert_to_column_letter(column_number):
    column_letter = ''
    while column_number != 0:
        c = ((column_number - 1) % 26)
        column_letter = chr(c + 65) + column_letter
        column_number = (column_number - c) // 26
    return column_letter


def update_table_article(name, last_monday, last_sunday, month, table_id, list_articles):
    range_name = f'{name} {last_monday}-{last_sunday}.{month}'
    position_for_place = START_POSITION_FOR_PLACE
    service = build('sheets', 'v4', credentials=credentials)
    sheet = service.spreadsheets()
    result = sheet.values().get(spreadsheetId=table_id,
                                range=range_name, majorDimension='ROWS').execute()
    values = result.get('values', [])
    i = 2
    body_data = []
    if not values:
        logging.info('No data found.')
    else:
        for row in values[1:]:
            try:
                index = 0
                if len(list_articles) != 0:
                    value = list_articles[index]
                    body_data += [
                        {'range': f'{range_name}!{convert_to_column_letter(position_for_place + 1)}{i}',
                         'values': [[f'{value}']]}]
                    list_articles.pop(index)
                    index += 1
            except:
                pass
            finally:
                i += 1
                body = {
                    'valueInputOption': 'USER_ENTERED',
                    'data': body_data}
    sheet.values().batchUpdate(spreadsheetId=table_id, body=body).execute()


def update_table_count(name, last_monday, last_sunday, month, table_id, dict_count):
    range_name = f'{name} {last_monday}-{last_sunday}.{month}'
    position_for_place = START_POSITION_FOR_PLACE

    service = build('sheets', 'v4', credentials=credentials)
    sheet = service.spreadsheets()
    result = sheet.values().get(spreadsheetId=table_id,
                                range=range_name, majorDimension='ROWS').execute()
    values = result.get('values', [])
    i = 2
    body_data = []
    if not values:
        logging.info('No data found.')
    else:
        for row in values[1:]:
            try:
                count = row[2].strip()
                if count in dict_count:
                    value = dict_count[count]
                    body_data += [
                        {'range': f'{range_name}!{convert_to_column_letter(position_for_place + 7)}{i}',
                         'values': [[f'{value}']]}]
            except:
                pass
            finally:
                i += 1
                body = {
                    'valueInputOption': 'USER_ENTERED',
                    'data': body_data}
    sheet.values().batchUpdate(spreadsheetId=table_id, body=body).execute()


def update_table_count_refund(name, last_monday, last_sunday, month, table_id, dict_count_refund):
    range_name = f'{name} {last_monday}-{last_sunday}.{month}'
    position_for_place = START_POSITION_FOR_PLACE

    service = build('sheets', 'v4', credentials=credentials)
    sheet = service.spreadsheets()
    result = sheet.values().get(spreadsheetId=table_id,
                                range=range_name, majorDimension='ROWS').execute()
    values = result.get('values', [])
    i = 2
    body_data = []
    if not values:
        logging.info('No data found.')
    else:
        for row in values[1:]:
            try:
                count = row[2].strip()
                if count in dict_count_refund:
                    value = dict_count_refund[count]
                    body_data += [
                        {'range': f'{range_name}!{convert_to_column_letter(position_for_place + 8)}{i}',
                         'values': [[f'{value}']]}]
            except:
                pass
            finally:
                i += 1
                body = {
                    'valueInputOption': 'USER_ENTERED',
                    'data': body_data}
    sheet.values().batchUpdate(spreadsheetId=table_id, body=body).execute()


def update_table_WB_sell(name, last_monday, last_sunday, month, table_id, dict_WB_sell):
    range_name = f'{name} {last_monday}-{last_sunday}.{month}'
    position_for_place = START_POSITION_FOR_PLACE

    service = build('sheets', 'v4', credentials=credentials)
    sheet = service.spreadsheets()
    result = sheet.values().get(spreadsheetId=table_id,
                                range=range_name, majorDimension='ROWS').execute()
    values = result.get('values', [])
    i = 2
    body_data = []
    if not values:
        logging.info('No data found.')
    else:
        for row in values[1:]:
            try:
                count = row[2].strip()
                if count in dict_WB_sell:
                    value = dict_WB_sell[count]
                    body_data += [
                        {'range': f'{range_name}!{convert_to_column_letter(position_for_place + 10)}{i}',
                         'values': [[f'{int(value)}']]}]
            except:
                pass
            finally:
                i += 1
                body = {
                    'valueInputOption': 'USER_ENTERED',
                    'data': body_data}
    sheet.values().batchUpdate(spreadsheetId=table_id, body=body).execute()


def update_table_logistics(name, last_monday, last_sunday, month, table_id, dict_logistics):
    range_name = f'{name} {last_monday}-{last_sunday}.{month}'
    position_for_place = START_POSITION_FOR_PLACE

    service = build('sheets', 'v4', credentials=credentials)
    sheet = service.spreadsheets()
    result = sheet.values().get(spreadsheetId=table_id,
                                range=range_name, majorDimension='ROWS').execute()
    values = result.get('values', [])
    i = 2
    body_data = []
    if not values:
        logging.info('No data found.')
    else:
        for row in values[1:]:
            try:
                count = row[2].strip()
                if count in dict_logistics:
                    value = dict_logistics[count]
                    body_data += [
                        {'range': f'{range_name}!{convert_to_column_letter(position_for_place + 12)}{i}',
                         'values': [[f'{int(value)}']]}]
            except:
                pass
            finally:
                i += 1
                body = {
                    'valueInputOption': 'USER_ENTERED',
                    'data': body_data}
    sheet.values().batchUpdate(spreadsheetId=table_id, body=body).execute()


def update_table_get_seller(name, last_monday, last_sunday, month, table_id, dict_get_seller):
    range_name = f'{name} {last_monday}-{last_sunday}.{month}'
    position_for_place = START_POSITION_FOR_PLACE

    service = build('sheets', 'v4', credentials=credentials)
    sheet = service.spreadsheets()
    result = sheet.values().get(spreadsheetId=table_id,
                                range=range_name, majorDimension='ROWS').execute()
    values = result.get('values', [])
    i = 2
    body_data = []
    if not values:
        logging.info('No data found.')
    else:
        for row in values[1:]:
            try:
                count = row[2].strip()
                if count in dict_get_seller:
                    value = dict_get_seller[count]
                    body_data += [
                        {'range': f'{range_name}!{convert_to_column_letter(position_for_place + 14)}{i}',
                         'values': [[f'{int(value)}']]}]
            except:
                pass
            finally:
                i += 1
                body = {
                    'valueInputOption': 'USER_ENTERED',
                    'data': body_data}
    sheet.values().batchUpdate(spreadsheetId=table_id, body=body).execute()


if __name__ == '__main__':
    table_id = SPREADSHEET_ID
    date_from = dt.datetime.date(dt.datetime.now()) - dt.timedelta(days=7)
    date_to = dt.datetime.date(dt.datetime.now()) - dt.timedelta(days=1)
    last_monday = date_from.strftime("%d")
    last_sunday = date_to.strftime("%d")
    month = date_to.strftime("%m")


    cred_file = os.path.join(BASE_DIR, 'credentials.json')
    with open(cred_file, 'r', encoding="utf-8") as f:
        cred = json.load(f)

    for i in cred:
        if i != 'Савельева':
            table_id = SPREADSHEET_ID
            name = i
            with warnings.catch_warnings(record=True):
                warnings.simplefilter("always")
                excel_file = openpyxl.load_workbook(f'excel_docs/{name} {last_monday}-{last_sunday}.{month}.xlsx')
            employees_sheet = excel_file['Sheet1']
            sell = dict_get_seller(employees_sheet)

            list_articles(employees_sheet)
            update_table_article(name, last_monday, last_sunday, month, table_id, list_articles(employees_sheet))
            update_table_count(name, last_monday, last_sunday, month, table_id, dict_count(employees_sheet))
            update_table_count_refund(name, last_monday, last_sunday, month, table_id,
                                      dict_count_refund(employees_sheet))
            update_table_WB_sell(name, last_monday, last_sunday, month, table_id, dict_WB_sell(employees_sheet))
            update_table_logistics(name, last_monday, last_sunday, month, table_id, dict_logistics(employees_sheet))
            update_table_get_seller(name, last_monday, last_sunday, month, table_id, dict_get_seller(employees_sheet))
